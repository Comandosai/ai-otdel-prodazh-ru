# -*- coding: utf-8 -*-
"""
fns_rmsp.py — выборка компаний из Единого реестра МСП (rmsp.nalog.ru).

Это ЯДРО конвейера. Один POST отдаёт всю нишу целиком в xlsx за секунды.
Капчи в этом сервисе нет, ключ не нужен, работает с любого IP.

Что реально приходит (проверено живыми запросами 13.08.2026):

  search-proc.json  — JSON, 10 записей на страницу, зато с полями
                      od2_sschr (численность) и has_contracts (госконтракты).
  report.xlsx       — вся выборка одним файлом (5084 строки за 2.0 с),
                      НО колонки «есть госконтракты» в файле НЕТ.
                      В xlsx 22 колонки: № п/п, Наименование, Тип субъекта,
                      Категория, ОГРН, ИНН, Основной вид деятельности, Регион,
                      Район, Город, Населенный пункт, Вновь созданный,
                      Дата включения, Дата исключения, Телефон, E-mail, WWW,
                      Лицензии, Инновационная продукция, Партнерство,
                      Социальное предприятие, Среднесписочная численность.

Отсюда рабочее правило конвейера:
  массовый список берём из report.xlsx, а флаг госконтрактов, если он нужен,
  добираем точечно через search() по конкретному ИНН.

Требования: Python 3.9, стандартная библиотека. Для чтения xlsx используется
openpyxl (см. requirements.txt), при его отсутствии включается встроенный
запасной парсер на zipfile + ElementTree.
"""

import io
import json
import os
import re
import socket
import ssl
import sys
import time
import zipfile
import datetime
import http.cookiejar
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET

from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

BASE = "https://rmsp.nalog.ru"
URL_SEARCH_HTML = BASE + "/search.html?mode=extended"
URL_SEARCH_JSON = BASE + "/search-proc.json"
URL_REPORT_XLSX = BASE + "/report.xlsx"

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")

TIMEOUT = 60
RETRIES = 3
THROTTLE_SEC = 0.5          # вежливый темп, сервис лимитов не показывал

_last_call = [0.0]


class RmspError(Exception):
    """Любая понятная человеку ошибка при работе с реестром МСП."""
    pass


# --------------------------------------------------------------------------
# сеть
# --------------------------------------------------------------------------

def get_cookies() -> urllib.request.OpenerDirector:
    """
    Забирает стартовые куки с формы расширенного поиска.

    Без этого шага search-proc.json и report.xlsx отвечают иначе.
    Возвращает opener с cookie-jar, его надо передавать во все остальные вызовы.
    """
    jar = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(jar))
    opener.addheaders = []
    req = urllib.request.Request(URL_SEARCH_HTML, headers={"User-Agent": UA})
    resp = _open(opener, req)
    resp.read()
    if len(jar) == 0:
        # не критично: сервис иногда отдаёт форму без Set-Cookie
        print("  внимание: сервер не выдал куки, продолжаю без них", file=sys.stderr)
    return opener


def _throttle() -> None:
    delta = time.time() - _last_call[0]
    if delta < THROTTLE_SEC:
        time.sleep(THROTTLE_SEC - delta)
    _last_call[0] = time.time()


def _open(opener: urllib.request.OpenerDirector,
          req: urllib.request.Request,
          timeout: int = TIMEOUT,
          retries: int = RETRIES):
    """Открывает запрос с ретраями. Все сообщения об ошибках по-русски."""
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            return opener.open(req, timeout=timeout)
        except urllib.error.HTTPError as err:
            body = ""
            try:
                body = err.read().decode("utf-8", "replace")[:400]
            except Exception:
                pass
            if err.code >= 500 and attempt < retries:
                last_err = err
                time.sleep(1.5 * attempt)
                continue
            raise RmspError(
                "Реестр МСП ответил HTTP %d на %s.\n  Ответ сервера: %s"
                % (err.code, req.full_url, body or "(пусто)"))
        except (urllib.error.URLError, socket.timeout, ssl.SSLError, OSError) as err:
            last_err = err
            if attempt < retries:
                time.sleep(1.5 * attempt)
                continue
    raise RmspError(
        "Не смог достучаться до rmsp.nalog.ru за %d попытки. Причина: %s.\n"
        "  Проверьте интернет, VPN и то, что домен не заблокирован."
        % (retries, last_err))


# --------------------------------------------------------------------------
# параметры формы
# --------------------------------------------------------------------------

def registry_date(today: Optional[datetime.date] = None) -> str:
    """
    Возвращает дату состояния реестра в формате 10.08.2026.

    Реестр МСП обновляется 10-го числа каждого месяца. Параметр dtCategory
    ОБЯЗАТЕЛЕН, и в нём должно стоять 10-е число уже вышедшего среза:
    если сегодня 10-е или позже, берём текущий месяц, иначе предыдущий.
    """
    if today is None:
        today = datetime.date.today()
    if today.day >= 10:
        d = datetime.date(today.year, today.month, 10)
    else:
        first = datetime.date(today.year, today.month, 1)
        prev = first - datetime.timedelta(days=1)
        d = datetime.date(prev.year, prev.month, 10)
    return d.strftime("%d.%m.%Y")


def _as_list(value: Union[None, str, int, Sequence]) -> List[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(v).strip() for v in value if str(v).strip()]
    return [p.strip() for p in str(value).split(",") if p.strip()]


def build_form(okved: Union[None, str, Sequence] = None,
               region: Union[None, str, int, Sequence] = None,
               np_type: Union[None, str, Sequence] = "UL",
               is_new: Optional[Union[bool, int]] = None,
               category: Union[None, int, str, Sequence] = None,
               dt_category: Optional[str] = None,
               query: Optional[str] = None,
               city: Optional[str] = None,
               district: Optional[str] = None,
               is_social: Optional[Union[bool, int]] = None,
               has_licenses: Optional[Union[bool, int]] = None,
               page_token: Optional[str] = None,
               sort_field: str = "NAME_EX",
               sort: str = "ASC") -> List[Tuple[str, str]]:
    """
    Собирает пары параметров формы ровно в том виде, в каком их ждёт сервис.

    okved  — код с точкой: 10.71, можно список ['10.71', '10.72'].
    region — код субъекта РФ по классификатору ФНС: 16 Татарстан, 77 Москва.
    np_type— UL юрлица, IP индивидуальные предприниматели.
    category — 1 микро, 2 малое, 3 среднее.
    page_token — ПОДПИСАННЫЙ токен из pageNav, а не номер страницы.
    """
    form: List[Tuple[str, str]] = [("mode", "extended")]

    if query:
        form.append(("query", query))
    for value in _as_list(np_type):
        form.append(("npType", value))
    for value in _as_list(category):
        form.append(("category", value))

    # dtCategory обязателен, иначе выборка ведёт себя непредсказуемо
    form.append(("dtCategory", dt_category or registry_date()))

    if is_new is not None:
        form.append(("isNew", "1" if is_new else "0"))
    if is_social is not None:
        form.append(("isSocial", "1" if is_social else "0"))
    if has_licenses is not None:
        form.append(("hasLicenses", "1" if has_licenses else "0"))

    for value in _as_list(region):
        form.append(("region", value))
    if district:
        form.append(("district", district))
    if city:
        form.append(("city", city))

    okveds = _as_list(okved)
    if okveds:
        # сервис принимает список через запятую одним значением
        form.append(("okved1", ",".join(okveds)))

    if page_token:
        form.append(("page", page_token))

    form.append(("sortField", sort_field))
    form.append(("sort", sort))
    return form


def _post(opener, url: str, form: List[Tuple[str, str]], timeout: int = TIMEOUT):
    _throttle()
    data = urllib.parse.urlencode(form).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={
            "User-Agent": UA,
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": URL_SEARCH_HTML,
            "Accept": "*/*",
        })
    return _open(opener, req, timeout=timeout)


# --------------------------------------------------------------------------
# JSON-поиск
# --------------------------------------------------------------------------

def search_raw(opener: Optional[urllib.request.OpenerDirector] = None,
               **kwargs) -> Dict[str, Any]:
    """
    Один POST на search-proc.json. Отдаёт полный JSON:
    rowCount, pageNav (токены страниц) и data (до 10 записей).
    """
    own = opener is None
    if own:
        opener = get_cookies()
    form = build_form(**kwargs)
    resp = _post(opener, URL_SEARCH_JSON, form)
    raw = resp.read().decode("utf-8", "replace")
    try:
        payload = json.loads(raw)
    except ValueError:
        raise RmspError(
            "Реестр МСП вернул не JSON. Первые 300 символов ответа:\n  %s"
            % raw[:300])
    if isinstance(payload, dict) and payload.get("ERRORS"):
        raise RmspError(
            "Реестр МСП отверг параметры запроса: %s"
            % json.dumps(payload["ERRORS"], ensure_ascii=False))
    return payload


def _next_page_token(payload: Dict[str, Any]) -> Optional[str]:
    nav = (payload.get("pageNav") or {}).get("pages") or []
    for item in nav:
        if not isinstance(item, (list, tuple)) or len(item) < 2:
            continue
        label = str(item[0]).strip().lower()
        if label in ("дальше", "далее", "вперед", "вперёд") and item[1]:
            return item[1]
    return None


def search(okved: Union[None, str, Sequence] = None,
           region: Union[None, str, int, Sequence] = None,
           np_type: Union[None, str, Sequence] = "UL",
           is_new: Optional[Union[bool, int]] = None,
           category: Union[None, int, str, Sequence] = None,
           max_pages: int = 1,
           opener: Optional[urllib.request.OpenerDirector] = None,
           **extra) -> List[Dict[str, Any]]:
    """
    Постраничный поиск. Возвращает список записей (по 10 на страницу).

    Это единственный путь, который отдаёт has_contracts (признак госконтрактов)
    и od2_sschr (численность) в машиночитаемом виде. В report.xlsx численность
    есть, а госконтрактов НЕТ.

    max_pages ограничивает глубину: 10 записей на страницу, поэтому для
    массовой выборки используйте download_report_xlsx, а search — для
    точечного добора по ИНН и для быстрого подсчёта размера ниши.

    Фильтр category сервером в одном из замеров был проигнорирован, поэтому
    результат дополнительно фильтруется локально по полю category.
    """
    own = opener is None
    if own:
        opener = get_cookies()

    wanted_categories = set(int(c) for c in _as_list(category)) if category else set()
    rows: List[Dict[str, Any]] = []
    token = None

    for page_no in range(max_pages):
        payload = search_raw(opener=opener, okved=okved, region=region,
                             np_type=np_type, is_new=is_new, category=category,
                             page_token=token, **extra)
        chunk = payload.get("data") or []
        rows.extend(chunk)
        token = _next_page_token(payload)
        if not token or not chunk:
            break

    if wanted_categories:
        rows = [r for r in rows if r.get("category") in wanted_categories]
    return rows


def count(okved: Union[None, str, Sequence] = None,
          region: Union[None, str, int, Sequence] = None,
          np_type: Union[None, str, Sequence] = "UL",
          opener: Optional[urllib.request.OpenerDirector] = None,
          **extra) -> int:
    """Быстрый ответ на вопрос «сколько всего в нише», один запрос."""
    payload = search_raw(opener=opener, okved=okved, region=region,
                         np_type=np_type, **extra)
    return int(payload.get("rowCount") or 0)


# --------------------------------------------------------------------------
# выгрузка всей выборки одним файлом
# --------------------------------------------------------------------------

def download_report_xlsx(okved: Union[None, str, Sequence] = None,
                         region: Union[None, str, int, Sequence] = None,
                         np_type: Union[None, str, Sequence] = "UL",
                         is_new: Optional[Union[bool, int]] = None,
                         category: Union[None, int, str, Sequence] = None,
                         out_path: Optional[str] = None,
                         opener: Optional[urllib.request.OpenerDirector] = None,
                         **extra) -> str:
    """
    Скачивает всю выборку одним xlsx и возвращает путь к файлу.

    Замеры: 290 записей за 1.1 с (48 КБ), 5084 записи за 2.0 с (698 КБ).
    Верхний предел размера выборки не проверялся, на очень широких запросах
    поведение неизвестно.
    """
    own = opener is None
    if own:
        opener = get_cookies()

    form = build_form(okved=okved, region=region, np_type=np_type,
                      is_new=is_new, category=category, **extra)
    resp = _post(opener, URL_REPORT_XLSX, form, timeout=180)
    blob = resp.read()

    if not blob[:2] == b"PK":
        head = blob[:300].decode("utf-8", "replace")
        raise RmspError(
            "Вместо xlsx пришло что-то другое (%d байт). Начало ответа:\n  %s"
            % (len(blob), head))

    if out_path is None:
        stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        tag = "-".join(filter(None, [
            "reg" + "_".join(_as_list(region)) if region else "",
            "okved" + "_".join(_as_list(okved)).replace(".", "") if okved else "",
        ])) or "rmsp"
        out_path = os.path.join(os.getcwd(), "rmsp-%s-%s.xlsx" % (tag, stamp))

    directory = os.path.dirname(os.path.abspath(out_path))
    if directory and not os.path.isdir(directory):
        os.makedirs(directory, exist_ok=True)
    with open(out_path, "wb") as fh:
        fh.write(blob)
    return out_path


# --------------------------------------------------------------------------
# разбор xlsx
# --------------------------------------------------------------------------

_COL_MAP = [
    (("п/п",), "num"),
    (("наимен", "фио"), "name"),
    (("тип субъект",), "np_type"),
    (("категор",), "category_text"),
    (("огрн",), "ogrn"),
    (("инн",), "inn"),
    (("основной вид", "вид деятельн"), "okved_text"),
    (("регион", "субъект рф"), "region_name"),
    (("район",), "district"),
    (("город",), "city"),
    (("населен",), "locality"),
    (("вновь",), "is_new"),
    (("дата включен",), "dt_registry"),
    (("дата исключен",), "dt_registry_out"),
    (("телефон",), "phone"),
    (("mail", "почт"), "email"),
    (("www", "сайт"), "site"),
    (("лицензи",), "has_licenses"),
    (("инновацион",), "is_hitech"),
    (("партнер", "партнёр"), "is_partnership"),
    (("социальн",), "is_social"),
    (("численност",), "employees"),
]

# В xlsx категория приходит текстом. Значение «Не является субъектом МСП»
# реально встречается: отчёт включает и тех, кого из реестра уже исключили,
# у таких строк заполнена «Дата исключения из реестра».
_CATEGORY_CODES = [("не является", 0), ("микро", 1), ("мал", 2), ("сред", 3)]


def _col_index(ref: str) -> int:
    """A -> 0, B -> 1, AA -> 26. Из ссылки вида 'AB12' берёт только буквы."""
    idx = 0
    for ch in ref:
        if ch.isalpha():
            idx = idx * 26 + (ord(ch.upper()) - 64)
        else:
            break
    return idx - 1


def _read_rows_stdlib(path: str) -> List[List[Any]]:
    """Запасной парсер xlsx без внешних библиотек: zipfile + ElementTree."""
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()

        shared: List[str] = []
        if "xl/sharedStrings.xml" in names:
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall(ns + "si"):
                shared.append("".join(t.text or "" for t in si.iter(ns + "t")))

        sheet = "xl/worksheets/sheet1.xml"
        if sheet not in names:
            candidates = sorted(n for n in names
                                if n.startswith("xl/worksheets/") and n.endswith(".xml"))
            if not candidates:
                raise RmspError("В файле %s нет ни одного листа" % path)
            sheet = candidates[0]

        root = ET.fromstring(zf.read(sheet))
        rows: List[List[Any]] = []
        for row_el in root.iter(ns + "row"):
            cells: Dict[int, Any] = {}
            max_idx = -1
            for cell in row_el.findall(ns + "c"):
                idx = _col_index(cell.get("r") or "")
                if idx < 0:
                    continue
                ctype = cell.get("t")
                if ctype == "inlineStr":
                    is_el = cell.find(ns + "is")
                    value = "".join(t.text or "" for t in is_el.iter(ns + "t")) \
                        if is_el is not None else None
                else:
                    v_el = cell.find(ns + "v")
                    value = v_el.text if v_el is not None else None
                    if ctype == "s" and value is not None:
                        pos = int(value)
                        value = shared[pos] if 0 <= pos < len(shared) else None
                cells[idx] = value
                if idx > max_idx:
                    max_idx = idx
            rows.append([cells.get(i) for i in range(max_idx + 1)])
        return rows


def _read_rows_openpyxl(path: str) -> List[List[Any]]:
    """
    Чтение через openpyxl. Нужен обход бага генератора ФНС.

    В файле от rmsp.nalog.ru стоит заведомо неверный <dimension ref="A1"/>,
    то есть «на листе одна ячейка». В режиме read_only openpyxl доверяет
    этому полю и молча отдаёт ОДНУ строку вместо 292. Лечится вызовом
    reset_dimensions(), который заставляет пересчитать границы по данным.
    """
    import openpyxl  # noqa
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        if hasattr(ws, "reset_dimensions"):
            ws.reset_dimensions()
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_rows(path: str) -> List[List[Any]]:
    """
    Основной путь чтения, намеренно без внешних зависимостей.

    Встроенный парсер на zipfile + ElementTree проверен на живом файле ФНС
    и оказался НАДЁЖНЕЕ openpyxl (см. историю с dimension выше), поэтому он
    идёт первым. openpyxl остаётся страховкой, если формат вдруг изменится.
    """
    try:
        rows = _read_rows_stdlib(path)
        if len(rows) >= 2:
            return rows
    except Exception as exc:                      # noqa: BLE001
        print("  встроенный парсер xlsx не справился (%s), пробую openpyxl"
              % exc, file=sys.stderr)
    try:
        return _read_rows_openpyxl(path)
    except ImportError:
        raise RmspError(
            "Не удалось прочитать %s встроенным парсером, а openpyxl не "
            "установлен.\n  Установите его: pip install -r requirements.txt" % path)


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _yesno(value: Any) -> Optional[int]:
    text = _norm(value).lower()
    if not text:
        return None
    if text.startswith("да") or text in ("1", "+", "true"):
        return 1
    if text.startswith("нет") or text in ("0", "-", "false"):
        return 0
    return None


def _to_int(value: Any) -> Optional[int]:
    text = _norm(value).replace(" ", "").replace(" ", "")
    if not text:
        return None
    match = re.search(r"-?\d+", text)
    return int(match.group(0)) if match else None


def _category_code(text: Any) -> Optional[int]:
    low = _norm(text).lower()
    if not low:
        return None
    for needle, code in _CATEGORY_CODES:
        if needle in low:
            return code
    return _to_int(low)


def _split_okved(text: Any) -> Tuple[Optional[str], Optional[str]]:
    """'10.71.1 Производство хлеба...' -> ('10.71.1', 'Производство хлеба...')"""
    raw = _norm(text)
    if not raw:
        return None, None
    match = re.match(r"^\s*(\d{2}(?:\.\d+)*)\s*[-–—:]?\s*(.*)$", raw)
    if match:
        return match.group(1), (match.group(2) or None)
    return None, raw


def _split_region(text: Any) -> Tuple[Optional[str], Optional[str]]:
    """'16 - Республика Татарстан (Татарстан)' -> ('16', 'Республика Татарстан...')"""
    raw = _norm(text)
    if not raw:
        return None, None
    match = re.match(r"^\s*(\d{1,2})\s*[-–—]\s*(.*)$", raw)
    if match:
        return match.group(1), (match.group(2) or None)
    return None, raw


def parse_report_xlsx(path: str) -> List[Dict[str, Any]]:
    """
    Читает скачанный report.xlsx и возвращает список словарей.

    Шапка в файле занимает две строки, поэтому колонки ищутся не по позиции,
    а по тексту заголовка: это переживёт добавление или перестановку колонок.

    Ключи результата: inn, ogrn, name, np_type, category (0 не МСП, 1 микро,
    2 малое, 3 среднее), category_text, okved, okved_name, region, region_name,
    district, city, locality, is_new, dt_registry, dt_registry_out, phone,
    email, site, has_licenses, is_hitech, is_partnership, is_social, employees.

    Два предупреждения по реальным данным:
      1. Колонки has_contracts в xlsx НЕТ, её надо добирать через search().
      2. В отчёт попадают и компании, ИСКЛЮЧЁННЫЕ из реестра: у них заполнена
         dt_registry_out и category = 0. Отсеивайте их через filter_records
         с active_only=True, иначе будете писать письма в никуда.
    """
    rows = _read_rows(path)
    if not rows:
        raise RmspError("Файл %s пустой" % path)

    header_idx = None
    for i, row in enumerate(rows[:15]):
        joined = " | ".join(_norm(c).lower() for c in row)
        if "инн" in joined and "огрн" in joined:
            header_idx = i
            break
    if header_idx is None:
        raise RmspError(
            "В файле %s не нашлась строка заголовка с колонками ИНН и ОГРН.\n"
            "  Похоже, сервис изменил формат отчёта." % path)

    header = rows[header_idx]
    mapping: Dict[int, str] = {}
    for col, cell in enumerate(header):
        low = _norm(cell).lower()
        if not low:
            continue
        for needles, key in _COL_MAP:
            if any(n in low for n in needles):
                if key not in mapping.values():
                    mapping[col] = key
                break

    if "inn" not in mapping.values():
        raise RmspError("В шапке файла %s не опознана колонка ИНН" % path)

    records: List[Dict[str, Any]] = []
    for row in rows[header_idx + 1:]:
        raw = {}
        for col, key in mapping.items():
            raw[key] = row[col] if col < len(row) else None

        inn = _norm(raw.get("inn"))
        if not inn or not inn.isdigit():
            continue

        okved_code, okved_name = _split_okved(raw.get("okved_text"))
        region_code, region_name = _split_region(raw.get("region_name"))
        records.append({
            "inn": inn,
            "ogrn": _norm(raw.get("ogrn")) or None,
            "name": _norm(raw.get("name")) or None,
            "np_type": _norm(raw.get("np_type")) or None,
            "category": _category_code(raw.get("category_text")),
            "category_text": _norm(raw.get("category_text")) or None,
            "okved": okved_code,
            "okved_name": okved_name,
            "region": region_code,
            "region_name": region_name,
            "district": _norm(raw.get("district")) or None,
            "city": _norm(raw.get("city")) or None,
            "locality": _norm(raw.get("locality")) or None,
            "is_new": _yesno(raw.get("is_new")),
            "dt_registry": _norm(raw.get("dt_registry")) or None,
            "dt_registry_out": _norm(raw.get("dt_registry_out")) or None,
            "phone": _norm(raw.get("phone")) or None,
            "email": _norm(raw.get("email")) or None,
            "site": _norm(raw.get("site")) or None,
            "has_licenses": _yesno(raw.get("has_licenses")),
            "is_hitech": _yesno(raw.get("is_hitech")),
            "is_partnership": _yesno(raw.get("is_partnership")),
            "is_social": _yesno(raw.get("is_social")),
            "employees": _to_int(raw.get("employees")),
        })
    return records


# --------------------------------------------------------------------------
# нормализация записей JSON-поиска
# --------------------------------------------------------------------------

def normalize_json_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Приводит запись из search-proc.json к тем же ключам, что и xlsx."""
    city = " ".join(filter(None, [_norm(row.get("citytype")),
                                  _norm(row.get("cityname"))])) or None
    district = " ".join(filter(None, [_norm(row.get("areaname")),
                                      _norm(row.get("areatype"))])) or None
    return {
        "inn": _norm(row.get("inn")) or None,
        "ogrn": _norm(row.get("ogrn")) or None,
        "name": _norm(row.get("name_ex")) or None,
        "np_type": _norm(row.get("nptype")) or None,
        "category": row.get("category"),
        "okved": _norm(row.get("okved1")) or None,
        "okved_name": _norm(row.get("okved1name")) or None,
        "region": _norm(row.get("regioncode")) or None,
        "district": district,
        "city": city,
        "employees": row.get("od2_sschr"),
        "has_contracts": row.get("has_contracts"),
        "has_licenses": row.get("has_licenses"),
        "is_new": row.get("isnew"),
        "is_active": row.get("is_active"),
        "dt_registry": _norm(row.get("dtregistry")) or None,
    }


# --------------------------------------------------------------------------
# отсев
# --------------------------------------------------------------------------

def filter_records(records: List[Dict[str, Any]],
                   exclude_okved_prefixes: Optional[Sequence[str]] = None,
                   include_okved_prefixes: Optional[Sequence[str]] = None,
                   employees_from: Optional[int] = None,
                   employees_to: Optional[int] = None,
                   categories: Optional[Sequence[int]] = None,
                   active_only: bool = False) -> List[Dict[str, Any]]:
    """
    Бесплатный отсев ДО обогащения, прямо на выгруженной таблице.

    Главный приём демо-ниши: транспортной компании не нужны клиенты, у которых
    есть свой автопарк. ОКВЭД 49.41 в карточке, значит перевозки они возят сами,
    и такую компанию надо выкинуть ещё до того, как мы потратили на неё
    хотя бы один сетевой запрос.

        filter_records(rows, exclude_okved_prefixes=['49.4'],
                       employees_from=15, active_only=True)

    active_only выкидывает исключённых из реестра: у них заполнена
    dt_registry_out либо категория равна 0.
    """
    out = []
    for rec in records:
        okved = rec.get("okved") or ""

        if active_only:
            if rec.get("dt_registry_out"):
                continue
            if rec.get("category") == 0:
                continue
            if rec.get("is_active") == 0:
                continue

        if include_okved_prefixes:
            if not any(okved.startswith(p) for p in include_okved_prefixes):
                continue
        if exclude_okved_prefixes:
            if any(okved.startswith(p) for p in exclude_okved_prefixes):
                continue

        emp = rec.get("employees")
        if employees_from is not None and (emp is None or emp < employees_from):
            continue
        if employees_to is not None and (emp is None or emp > employees_to):
            continue

        if categories and rec.get("category") not in set(categories):
            continue

        out.append(rec)
    return out


# --------------------------------------------------------------------------
# демо
# --------------------------------------------------------------------------

def _demo() -> None:
    region = "16"
    okved = "10.71"

    print("=" * 74)
    print("РЕЕСТР МСП. Демо-прогон: регион %s, ОКВЭД %s, юрлица" % (region, okved))
    print("Дата состояния реестра (dtCategory): %s" % registry_date())
    print("=" * 74)

    print("\n[1/4] Забираю куки с формы расширенного поиска")
    opener = get_cookies()
    print("      готово")

    print("\n[2/4] Считаю размер ниши через search-proc.json")
    started = time.time()
    payload = search_raw(opener=opener, okved=okved, region=region, np_type="UL")
    total = int(payload.get("rowCount") or 0)
    print("      найдено записей: %d (за %.1f с)" % (total, time.time() - started))

    rows = [normalize_json_row(r) for r in (payload.get("data") or [])]
    print("      первые записи (тут видно то, чего НЕТ в xlsx, "
          "признак госконтрактов):")
    for rec in rows[:3]:
        print("        %-12s %-46s чел: %-5s госконтракты: %s"
              % (rec["inn"],
                 (rec["name"] or "")[:46],
                 rec["employees"],
                 "да" if rec.get("has_contracts") else "нет"))

    print("\n[3/4] Скачиваю всю выборку одним xlsx")
    started = time.time()
    path = download_report_xlsx(okved=okved, region=region, np_type="UL",
                                out_path=os.path.join(
                                    os.getcwd(), "demo-rmsp-%s-%s.xlsx"
                                    % (region, okved.replace(".", ""))),
                                opener=opener)
    size = os.path.getsize(path)
    print("      %s" % path)
    print("      %d байт за %.1f с" % (size, time.time() - started))

    print("\n[4/4] Разбираю файл")
    records = parse_report_xlsx(path)
    print("      строк с данными: %d" % len(records))

    with_email = sum(1 for r in records if r.get("email"))
    with_phone = sum(1 for r in records if r.get("phone"))
    with_site = sum(1 for r in records if r.get("site"))
    with_emp = sum(1 for r in records if r.get("employees") is not None)
    print("      заполнено добровольно: почта %d, телефон %d, сайт %d"
          % (with_email, with_phone, with_site))
    print("      численность известна у %d из %d" % (with_emp, len(records)))

    excluded = sum(1 for r in records if r.get("dt_registry_out"))
    print("      из них уже ИСКЛЮЧЕНЫ из реестра: %d "
          "(у них заполнена дата исключения)" % excluded)

    print("\n      три первые компании из файла:")
    for rec in records[:3]:
        print("        %-12s %-38s %-9s %-5s чел.  %s"
              % (rec["inn"],
                 (rec["name"] or "")[:38],
                 rec["okved"],
                 rec["employees"],
                 rec["category_text"]))

    alive = filter_records(records, active_only=True)
    big = filter_records(records, active_only=True, employees_from=20,
                         exclude_okved_prefixes=["49.4"])
    print("\n      бесплатный отсев без единого лишнего запроса:")
    print("        всего в выгрузке          %d" % len(records))
    print("        живые в реестре           %d" % len(alive))
    print("        + от 20 человек, без своего автопарка (ОКВЭД 49.4)  %d" % len(big))

    print("\nГотово. Дальше эти ИНН идут в fns_bo.py за выручкой "
          "и в fns_egrul.py за ФИО директора.")


# --------------------------------------------------------------------------
# запись выборки в базу конвейера
# --------------------------------------------------------------------------

HERE = os.path.dirname(os.path.abspath(__file__))


def _verify():
    """Модуль verify лежит рядом. Импорт ленивый, чтобы --help ничего не грузил."""
    if HERE not in sys.path:
        sys.path.insert(0, HERE)
    import verify
    return verify


def _address(rec: Dict[str, Any]) -> Optional[str]:
    parts = [rec.get("region_name"), rec.get("district"), rec.get("city"),
             rec.get("locality")]
    seen, out = set(), []
    for part in parts:
        text = _norm(part)
        if text and text.lower() not in seen:
            seen.add(text.lower())
            out.append(text)
    return ", ".join(out) or None


def company_row(rec: Dict[str, Any], stamp: str) -> Dict[str, Any]:
    """Запись реестра в строку таблицы companies."""
    return {
        "inn": rec.get("inn"),
        "ogrn": rec.get("ogrn"),
        "name": rec.get("name"),
        "okved": rec.get("okved"),
        "okved_name": rec.get("okved_name"),
        "region": rec.get("region"),
        "city": rec.get("city"),
        "employees": rec.get("employees"),
        "category": rec.get("category"),
        "is_new": rec.get("is_new"),
        "has_contracts": rec.get("has_contracts"),
        "website": rec.get("site"),
        "email": rec.get("email"),
        "phone": rec.get("phone"),
        "address": _address(rec),
        "dt_registry": rec.get("dt_registry"),
        "collected_at": stamp,
    }


def save_to_db(records: List[Dict[str, Any]],
               db_path: str,
               exclude_inn: Optional[Sequence[str]] = None,
               verbose: bool = True) -> Dict[str, int]:
    """
    Кладёт разобранные записи в таблицу companies базы конвейера.

    Запись идёт через db_upsert_company, то есть по ИНН: повторный прогон
    обновляет строку, а не добавляет вторую. Уже собранные другими шагами
    поля (сайт, почта, выручка, директор) при этом не затираются пустотой.
    """
    verify = _verify()
    stop = set(verify.norm_digits(i) for i in (exclude_inn or []) if verify.norm_digits(i))
    con = verify.open_db(db_path)
    verify.ensure_company_columns(con)
    stamp = verify.now_stamp()
    written, skipped, bad = 0, 0, 0
    for rec in records:
        inn = verify.norm_digits(rec.get("inn"))
        if not inn:
            bad += 1
            continue
        if inn in stop:
            skipped += 1
            continue
        if verify.db_upsert_company(con, company_row(dict(rec, inn=inn), stamp)):
            written += 1
    con.commit()
    total = con.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    con.close()
    if verbose:
        print("      записано в базу: %d" % written)
        if skipped:
            print("      пропущено по стоп-листу isklyuchit_inn: %d" % skipped)
        if bad:
            print("      строк без ИНН: %d" % bad)
        print("      всего уникальных компаний в базе: %d" % total)
    return {"written": written, "skipped": skipped, "bad": bad, "total": total}


# --------------------------------------------------------------------------
# командная строка
# --------------------------------------------------------------------------

def _cli(argv: Sequence[str]) -> int:
    import argparse

    parser = argparse.ArgumentParser(
        prog="fns_rmsp.py",
        description="Выборка компаний из реестра МСП и запись её в базу конвейера",
        epilog="Без ключа --live скрипт в сеть не ходит и файлов не создаёт. "
               "Пример: fns_rmsp.py --config config.json --db data/leads.db "
               "--region 16 --okved 10.71 --limit 500 --live")
    parser.add_argument("--config", help="конфиг ниши, по умолчанию config.json "
                                         "в корне репозитория")
    parser.add_argument("--db", help="файл sqlite, по умолчанию data/leads.db")
    parser.add_argument("--region", help="код региона ФНС, например 16")
    parser.add_argument("--okved", help="коды ОКВЭД через запятую, например 10.71,10.51")
    parser.add_argument("--np-type", dest="np_type",
                        help="UL юрлица, IP предприниматели, по умолчанию UL")
    parser.add_argument("--category", help="категории МСП через запятую: 1 микро, "
                                           "2 малое, 3 среднее")
    parser.add_argument("--limit", type=int, help="сколько записей класть в базу")
    parser.add_argument("--employees-from", dest="employees_from", type=int)
    parser.add_argument("--employees-to", dest="employees_to", type=int)
    parser.add_argument("--exclude-okved", dest="exclude_okved",
                        help="какие ОКВЭД выкинуть, через запятую")
    parser.add_argument("--keep-excluded", action="store_true",
                        help="оставить тех, кого уже исключили из реестра МСП")
    parser.add_argument("--xlsx", help="файл выгрузки: с --live сюда сохраняем, "
                                       "с --offline отсюда читаем")
    parser.add_argument("--live", action="store_true",
                        help="разрешить обращение к rmsp.nalog.ru")
    parser.add_argument("--offline", action="store_true",
                        help="без сети: разобрать готовый --xlsx или просто "
                             "показать параметры выборки")
    parser.add_argument("--demo", action="store_true",
                        help="показательный прогон по Татарстану, требует --live")
    args = parser.parse_args(list(argv))

    verify = _verify()
    profile = verify.config_profile(verify.load_config(args.config))

    region = args.region or profile.get("region")
    okveds = verify.cfg_list(args.okved) or profile.get("okved") or []
    np_type = verify.cfg_list(args.np_type) or profile.get("np_type") or ["UL"]
    category = verify.cfg_list(args.category) or profile.get("category") or []
    employees_from = args.employees_from if args.employees_from is not None \
        else profile.get("employees_from")
    employees_to = args.employees_to if args.employees_to is not None \
        else profile.get("employees_to")
    exclude_okved = verify.cfg_list(args.exclude_okved) or profile.get("exclude_okved") or []
    db_path = verify.resolve_db_path(args.db, profile)

    print("Реестр МСП. Параметры выборки:")
    print("      конфиг     : %s" % (profile.get("path") or "не найден, беру ключи"))
    print("      регион     : %s" % (region or "любой"))
    print("      ОКВЭД      : %s" % (", ".join(okveds) or "любой"))
    print("      тип        : %s" % ", ".join(np_type))
    print("      численность: от %s до %s" % (employees_from or "любой",
                                              employees_to or "любой"))
    print("      база       : %s" % db_path)

    if args.demo:
        if not args.live:
            print("\nдемо ходит в сеть, поэтому запускается только с ключом --live")
            return 0
        _demo()
        return 0

    if not args.live and not args.offline:
        print("\nрежим не выбран. Добавьте --live, чтобы сходить в реестр, "
              "или --offline, чтобы работать без сети.")
        return 0

    if args.offline:
        if not args.xlsx:
            print("\nрежим --offline: сети нет, файл выгрузки не задан. "
                  "Передайте --xlsx с готовой выгрузкой, чтобы разобрать её в базу.")
            return 0
        if not os.path.exists(args.xlsx):
            print("\nфайла %s нет" % args.xlsx)
            return 1
        print("\n[1/2] Разбираю готовую выгрузку %s" % args.xlsx)
        path = args.xlsx
    else:
        path = args.xlsx
        if not path:
            stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
            folder = profile.get("xlsx_dir") or os.path.join("data", "xlsx")
            if not os.path.isabs(folder):
                folder = os.path.join(os.path.dirname(HERE), folder)
            path = os.path.join(folder, "rmsp-%s-%s-%s.xlsx"
                                % (region or "all",
                                   ("_".join(okveds).replace(".", "") or "all")[:40],
                                   stamp))
        print("\n[1/2] Скачиваю выгрузку из реестра МСП")
        started = time.time()
        path = download_report_xlsx(okved=okveds or None, region=region,
                                    np_type=np_type, category=category or None,
                                    dt_category=profile.get("dt_category"),
                                    out_path=path)
        print("      %s, %d байт за %.1f с"
              % (path, os.path.getsize(path), time.time() - started))

    records = parse_report_xlsx(path)
    print("      строк с данными: %d" % len(records))

    kept = filter_records(records,
                          exclude_okved_prefixes=exclude_okved or None,
                          employees_from=employees_from,
                          employees_to=employees_to,
                          categories=[int(c) for c in category if str(c).isdigit()] or None,
                          active_only=not args.keep_excluded)
    print("      после бесплатного отсева: %d" % len(kept))
    if args.limit:
        kept = kept[:args.limit]
        print("      ограничение --limit: %d" % len(kept))

    print("\n[2/2] Кладу в базу %s" % db_path)
    save_to_db(kept, db_path, exclude_inn=profile.get("exclude_inn"))
    print("\nГотово. Дальше выручка через fns_bo.py, ФИО директора через fns_egrul.py.")
    return 0


def main(argv: Optional[Sequence[str]] = None) -> int:
    argv = sys.argv[1:] if argv is None else list(argv)
    try:
        return _cli(argv)
    except RmspError as exc:
        print("\nОШИБКА: %s" % exc, file=sys.stderr)
        return 1
    except KeyboardInterrupt:
        print("\nПрервано пользователем", file=sys.stderr)
        return 130


if __name__ == "__main__":
    sys.exit(main())
